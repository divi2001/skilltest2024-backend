"""
Script to clear old subjectsdb and batchdb data and import new data from
subjectsdb.xlsx and batchdb.xlsx into the 'dec25' MySQL database.
Only updates columns that exist in the DB tables.
"""

import openpyxl
import mysql.connector
from datetime import datetime, time

# --- Config ---
SUBJECTSDB_EXCEL = r"C:\freelancer\kk exams software\subjectsdb.xlsx"
BATCHDB_EXCEL = r"C:\freelancer\kk exams software\batchdb.xlsx"
DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "root",
    "password": "tanuj1221",
    "database": "dec25"
}

# subjectsdb: only columns present in DB
SUBJECTSDB_COLUMNS = [
    "subjectId", "examType", "courseId", "subject_name", "subject_name_short",
    "daily_timer", "passage_timer", "demo_timer",
    "disability_passage_timer", "typing_timer", "disability_typing_timer"
]

# batchdb: only columns present in DB (skip gate_closure_time)
BATCHDB_COLUMNS = [
    "departmentId", "batchNo", "batchdate",
    "reporting_time", "start_time", "end_time", "batchstatus"
]


def parse_time_value(val):
    if val is None:
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S")
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ("%I:%M %p", "%H:%M:%S", "%H:%M", "%I:%M:%S %p"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%H:%M:%S")
        except ValueError:
            continue
    print(f"  WARNING: Could not parse time value: '{val_str}'")
    return None


def parse_date_value(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    print(f"  WARNING: Could not parse date value: '{val_str}'")
    return None


def parse_int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    val_str = str(val).strip()
    if not val_str:
        return None
    try:
        return int(float(val_str))
    except ValueError:
        return None


def parse_bool(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, (int, float)):
        return int(val)
    val_str = str(val).strip().lower()
    if val_str in ("1", "true", "yes"):
        return 1
    if val_str in ("0", "false", "no", ""):
        return 0
    return None


def import_table(cursor, conn, excel_path, table_name, db_columns, time_cols, date_cols, int_cols, bool_cols):
    print(f"\n{'='*50}")
    print(f"Importing {table_name} from {excel_path}")
    print(f"DB columns to import: {db_columns}")
    print(f"{'='*50}")

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    # Build header index
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {h: i for i, h in enumerate(headers)}
    print(f"Excel headers: {headers}")

    # Check which DB columns exist in Excel
    missing_in_excel = [c for c in db_columns if c not in header_idx]
    if missing_in_excel:
        print(f"WARNING: These DB columns are NOT in Excel (will be set to NULL): {missing_in_excel}")

    skipped_excel_cols = [h for h in headers if h not in db_columns]
    if skipped_excel_cols:
        print(f"Skipping Excel columns not in DB: {skipped_excel_cols}")

    # Read rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for col in db_columns:
            col_idx = header_idx.get(col)
            if col_idx is None:
                record[col] = None
                continue
            val = row[col_idx]

            if col in time_cols:
                record[col] = parse_time_value(val)
            elif col in date_cols:
                record[col] = parse_date_value(val)
            elif col in int_cols:
                record[col] = parse_int(val)
            elif col in bool_cols:
                record[col] = parse_bool(val)
            else:
                record[col] = str(val) if val is not None else None
        rows.append(record)

    wb.close()
    print(f"Loaded {len(rows)} rows from Excel")

    if not rows:
        print("No data to import. Skipping.")
        return

    # Current count
    cursor.execute(f"SELECT COUNT(*) FROM `{table_name}`")
    old_count = cursor.fetchone()[0]
    print(f"Current rows in {table_name}: {old_count}")

    # Delete old data
    print(f"Deleting all old records from {table_name}...")
    cursor.execute(f"DELETE FROM `{table_name}`")
    conn.commit()
    print(f"Deleted {old_count} old records.")

    # Insert new data
    placeholders = ", ".join(["%s"] * len(db_columns))
    col_names = ", ".join([f"`{c}`" for c in db_columns])
    insert_sql = f"INSERT INTO `{table_name}` ({col_names}) VALUES ({placeholders})"

    success = 0
    errors = 0
    for i, record in enumerate(rows):
        values = tuple(record[c] for c in db_columns)
        try:
            cursor.execute(insert_sql, values)
            success += 1
        except Exception as e:
            errors += 1
            if errors <= 10:
                print(f"  ERROR row {i+2}: {e}")
                print(f"    Values: {values}")

    conn.commit()

    # Verify
    cursor.execute(f"SELECT COUNT(*) FROM `{table_name}`")
    new_count = cursor.fetchone()[0]

    print(f"\n{table_name} IMPORT COMPLETE:")
    print(f"  Old records deleted: {old_count}")
    print(f"  New records inserted: {success}")
    print(f"  Errors: {errors}")
    print(f"  Current rows in DB: {new_count}")


def main():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # Disable FK checks so we can insert departmentIds not yet in departmentdb
    cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

    # Import subjectsdb
    import_table(
        cursor, conn,
        excel_path=SUBJECTSDB_EXCEL,
        table_name="subjectsdb",
        db_columns=SUBJECTSDB_COLUMNS,
        time_cols=[],
        date_cols=[],
        int_cols=["subjectId", "courseId", "daily_timer", "passage_timer", "demo_timer",
                  "disability_passage_timer", "typing_timer", "disability_typing_timer"],
        bool_cols=[]
    )

    # Import batchdb
    import_table(
        cursor, conn,
        excel_path=BATCHDB_EXCEL,
        table_name="batchdb",
        db_columns=BATCHDB_COLUMNS,
        time_cols=["reporting_time", "start_time", "end_time"],
        date_cols=["batchdate"],
        int_cols=["departmentId", "batchNo"],
        bool_cols=["batchstatus"]
    )

    cursor.close()
    conn.close()
    print(f"\n{'='*50}")
    print("ALL DONE!")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
