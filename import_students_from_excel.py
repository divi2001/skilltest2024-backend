"""
Script to clear old students data and import new data from combined_all.xlsx
into the 'students' table in the 'dec25' MySQL database.
"""

import openpyxl
import mysql.connector
from datetime import datetime, time

# --- Config ---
EXCEL_PATH = r"C:\freelancer\kk exams software\combined_all_dept.xlsx"
DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "root",
    "password": "tanuj1221",
    "database": "dec25"
}

# Mapping: students table column -> Excel header
COLUMN_MAP = {
    "student_id":    "student_id",
    "password":      "password",
    "instituteId":   None,            # not in Excel
    "batchNo":       "batchNo",
    "batchdate":     "batchdate",
    "fullname":      "fullname",
    "subjectsId":    "subjectsId",
    "courseId":       "courseId",
    "batch_year":    "batch_year",
    "loggedin":      "loggedin",
    "done":          "done",
    "photo":         "photo",
    "center":        "center",
    "reporting_time":"reporting_time",
    "start_time":    "start_time",
    "end_time":      "end_time",
    "day":           "day",
    "qset":          "qset",
    "base64":        "base64",
    "sign_base64":   "sign_base64",
    "IsShorthand":   "IsShorthand",
    "IsTypewriting": "IsTypewriting",
    "departmentId":  "departmentId",
    "disability":    "disability",
    "mothername":    "MOTHER NAME",
}

def parse_time_value(val):
    """Parse time from Excel - could be datetime.time, string like '10:00 AM', or None."""
    if val is None:
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S")
    # String like '10:00 AM' or '08:30 AM'
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ("%I:%M %p", "%H:%M:%S", "%H:%M", "%I:%M:%S %p"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%H:%M:%S")
        except ValueError:
            continue
    print(f"  WARNING: Could not parse time value: '{val_str}'")
    return None

def parse_date_value(val):
    """Parse date from Excel - could be datetime, string like '06-Mar-2026', or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    if not val_str:
        return None
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    print(f"  WARNING: Could not parse date value: '{val_str}'")
    return None

def parse_bool(val):
    """Parse boolean/tinyint value."""
    if val is None:
        return None
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, (int, float)):
        return int(val)
    val_str = str(val).strip().lower()
    if val_str in ("1", "true", "yes"):
        return 1
    if val_str in ("0", "false", "no", ""):
        return 0
    return None

def parse_int(val):
    """Parse integer value."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    val_str = str(val).strip()
    if not val_str:
        return None
    try:
        return int(float(val_str))
    except ValueError:
        return None

def main():
    # Load Excel
    print(f"Loading Excel file: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    
    # Build header index
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {h: i for i, h in enumerate(headers)}
    print(f"Found {len(headers)} columns, reading rows...")
    
    # Read all rows
    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Get student_id - required field
        sid_col = header_idx.get("student_id")
        if sid_col is None:
            print("ERROR: 'student_id' column not found in Excel!")
            return
        student_id = row[sid_col]
        if student_id is None:
            skipped += 1
            continue
        
        record = {}
        for db_col, excel_col in COLUMN_MAP.items():
            if excel_col is None:
                record[db_col] = None
                continue
            col_idx = header_idx.get(excel_col)
            if col_idx is None:
                record[db_col] = None
                continue
            val = row[col_idx]
            
            # Type-specific parsing
            if db_col == "student_id":
                record[db_col] = parse_int(val)
            elif db_col == "batchdate":
                record[db_col] = parse_date_value(val)
            elif db_col in ("reporting_time", "start_time", "end_time"):
                record[db_col] = parse_time_value(val)
            elif db_col in ("loggedin", "done", "IsShorthand", "IsTypewriting", "disability"):
                record[db_col] = parse_bool(val)
            elif db_col in ("batchNo", "subjectsId", "courseId", "center", "day", "qset", "departmentId", "instituteId"):
                record[db_col] = parse_int(val)
            else:
                record[db_col] = str(val) if val is not None else None
        
        if record["student_id"] is not None:
            rows.append(record)
    
    wb.close()
    print(f"Loaded {len(rows)} valid rows ({skipped} skipped due to missing student_id)")
    
    if not rows:
        print("No data to import. Exiting.")
        return
    
    # Connect to MySQL
    print(f"\nConnecting to MySQL ({DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']})...")
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    
    # Check current count
    cursor.execute("SELECT COUNT(*) FROM students")
    old_count = cursor.fetchone()[0]
    print(f"Current students in DB: {old_count}")
    
    # Clear old data
    print("Deleting all old student records...")
    cursor.execute("DELETE FROM students")
    conn.commit()
    print(f"Deleted {old_count} old records.")
    
    # Insert new data
    db_columns = list(COLUMN_MAP.keys())
    placeholders = ", ".join(["%s"] * len(db_columns))
    col_names = ", ".join([f"`{c}`" for c in db_columns])
    insert_sql = f"INSERT INTO students ({col_names}) VALUES ({placeholders})"
    
    success = 0
    errors = 0
    for i, record in enumerate(rows):
        values = tuple(record[c] for c in db_columns)
        try:
            cursor.execute(insert_sql, values)
            success += 1
        except Exception as e:
            errors += 1
            if errors <= 10:  # Show first 10 errors
                print(f"  ERROR row {i+2} (student_id={record.get('student_id')}): {e}")
    
    conn.commit()
    
    # Verify
    cursor.execute("SELECT COUNT(*) FROM students")
    new_count = cursor.fetchone()[0]
    
    print(f"\n{'='*50}")
    print(f"IMPORT COMPLETE")
    print(f"  Old records deleted: {old_count}")
    print(f"  New records inserted: {success}")
    print(f"  Errors: {errors}")
    print(f"  Current students in DB: {new_count}")
    print(f"{'='*50}")
    
    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
