"""
Audio Transcription Report Generator V2
========================================
Uses faster-whisper with large-v3 model for better Hindi/Marathi accuracy
This is a GPU-accelerated, open-source, free solution.
Generates Excel report with audiodb fields + transcriptions
"""

import os
import sys
import json
import mysql.connector
from datetime import datetime
from pydub import AudioSegment
from typing import Optional, Dict, List
import io
import warnings
import tempfile
import requests
warnings.filterwarnings("ignore")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import torch
    import torchaudio
    from transformers import AutoModel, AutoProcessor, pipeline
    TRANSFORMERS_AVAILABLE = True
except ImportError:
    TRANSFORMERS_AVAILABLE = False
    print("Transformers/Torch/Torchaudio not installed. Installing now...")

# Database configuration — reads from .env if available
import dotenv
dotenv.load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

DB_CONFIG = {
    'host': os.getenv('DB_HOST', '13.204.48.33').strip('"'),
    'port': int(os.getenv('DB_PORT', 3306)),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', 'tanuj1221'),
    'database': os.getenv('DB_DATABASE', 'skilmar26'),
    'charset': 'utf8mb4'
}

# Department filter — set to None to process all departments
TARGET_DEPARTMENT_ID = 13

# Audio processing configuration
# Audio processing configuration
AUDIO_DURATION_SECONDS = None  # Set to None for full audio
SAMPLE_RATE = 16000
HF_TOKEN = os.getenv("HF_TOKEN")

def install_dependencies():
    """Install required packages."""
    import subprocess
    print("Installing transformers torch torchaudio...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "transformers", "torch", "torchaudio", "accel-brain-base"])
    print("✓ Dependencies installed successfully")

def check_gpu_availability():
    """Check if GPU is available for processing."""
    import torch
    if torch.cuda.is_available():
        gpu_name = torch.cuda.get_device_name(0)
        gpu_memory = torch.cuda.get_device_properties(0).total_memory / (1024**3)
        print(f"✓ GPU Available: {gpu_name} ({gpu_memory:.2f} GB)")
        return True
    else:
        print("✗ GPU not available. Using CPU (slower).")
        return False

def load_indic_model():
    """Load AI4Bharat Indic Conformer model (for Hindi/Marathi)."""
    
    # Check GPU availability
    device = "cuda" if torch.cuda.is_available() else "cpu"
    
    print(f"Loading AI4Bharat Indic-Conformer-600M on {device.upper()} (for Hindi/Marathi)...")
    
    try:
        model = AutoModel.from_pretrained(
            "ai4bharat/indic-conformer-600m-multilingual", 
            trust_remote_code=True
        )
        model.to(device)
        model.eval()
        
        print(f"✓ Indic Model loaded successfully on {device.upper()}")
        return {
            'model': model,
            'device': device,
            'type': 'indic'
        }
    except Exception as e:
        print(f"✗ Failed to load Indic model: {e}")
        sys.exit(1)

def load_english_model():
    """Load English Model (Distil-Whisper Large-v2) - PyTorch."""
    device = "cuda" if torch.cuda.is_available() else "cpu"
    
    print(f"Loading English Model (Distil-Whisper Large-v2) on {device.upper()}...")
    try:
        # Use simple pipeline for speed and ease - strictly PyTorch
        # We explicitly set framework='pt' to avoid Keras checks if TF is installed
        pipe = pipeline(
            "automatic-speech-recognition",
            model="distil-whisper/distil-large-v2",
            dtype=torch.float16 if device == "cuda" else torch.float32,
            device=device,
            framework="pt"
        )
        print(f"✓ English Model loaded successfully")
        return {
            'model': pipe,
            'type': 'english-pipeline'
        }
    except Exception as e:
        print(f"✗ Failed to load English model: {e}")
        sys.exit(1)

def connect_to_database():
    """Establish connection to MySQL database."""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        print(f"✓ Connected to database: {DB_CONFIG['database']}")
        return connection
    except mysql.connector.Error as err:
        print(f"✗ Database connection failed: {err}")
        sys.exit(1)

def fetch_audio_records(connection, department_id=None) -> List[Dict]:
    """Fetch audio records from audiodb table, optionally filtered by department."""
    cursor = connection.cursor(dictionary=True)

    where_clause = "WHERE a.departmentId = %s" if department_id is not None else ""
    params = (department_id,) if department_id is not None else ()

    query = f"""
    SELECT 
        a.id, a.subjectId, a.qset, a.departmentId,
        a.code_a, a.code_b, a.code_t,
        a.audio1, a.audio2, a.testaudio,
        a.passage1, a.passage2,
        a.textPassageA, a.textPassageB,
        s.subject_name
    FROM audiodb a
    LEFT JOIN (
        SELECT subjectId, MAX(subject_name) as subject_name 
        FROM subjectsdb 
        GROUP BY subjectId
    ) s ON a.subjectId = s.subjectId
    {where_clause}
    ORDER BY a.subjectId, a.qset
    """
    cursor.execute(query, params)
    records = cursor.fetchall()
    cursor.close()

    dept_label = f" for departmentId={department_id}" if department_id else ""
    print(f"✓ Fetched {len(records)} audio records{dept_label}")
    return records

def detect_language_from_subject(subject_name: Optional[str]) -> str:
    """Detect language based on subject name."""
    if subject_name is None:
        return None
    
    subject_lower = subject_name.lower()
    
    if 'hindi' in subject_lower or 'हिंदी' in subject_lower:
        return 'hi'
    elif 'marathi' in subject_lower or 'मराठी' in subject_lower:
        return 'mr'
    elif 'english' in subject_lower:
        return 'en'
    else:
        return None  # Auto-detect

def download_audio(url: str, timeout: int = 60) -> Optional[bytes]:
    """Download audio file from URL."""
    if not url or url.strip() == '':
        return None
    
    try:
        response = requests.get(url, timeout=timeout, stream=True)
        response.raise_for_status()
        return response.content
    except requests.RequestException as e:
        print(f"  ⚠ Failed to download audio: {e}")
        return None

def format_duration(seconds: float) -> str:
    """Format duration in seconds to MM:SS or HH:MM:SS string."""
    seconds = int(seconds)
    if seconds >= 3600:
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    else:
        minutes = seconds // 60
        secs = seconds % 60
        return f"{minutes:02d}:{secs:02d}"

def extract_audio_content(audio_bytes: bytes, format: str = "mp3") -> tuple:
    """Extract audio content and save to temp file (supports full length).
    Returns (temp_file_path, duration_seconds) or (None, 0)."""
    try:
        # Load audio from bytes
        audio = AudioSegment.from_file(io.BytesIO(audio_bytes), format=format)
        
        # Get full duration in seconds before any truncation
        duration_seconds = len(audio) / 1000.0
        
        # Truncate if limit set
        if AUDIO_DURATION_SECONDS is not None:
             duration_ms = AUDIO_DURATION_SECONDS * 1000
             audio = audio[:duration_ms]
        
        # Convert to mono and set sample rate
        audio = audio.set_channels(1).set_frame_rate(SAMPLE_RATE)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
        audio.export(temp_file.name, format="wav")  # Ensure WAV for librosa if needed
        temp_file.close()
        
        return temp_file.name, duration_seconds
    except Exception as e:
        print(f"  ⚠ Failed to process audio: {e}")
        return None, 0

def transcribe_indic(model_dict, audio_path: str, language: Optional[str] = None) -> Dict:
    """Transcribe audio using Indic Conformer model."""
    try:
        model = model_dict['model']
        device = model_dict['device']
        
        # Load audio using torchaudio
        wav, sr = torchaudio.load(audio_path)
        
        # Resample if needed (target 16000)
        target_sr = 16000
        if sr != target_sr:
            resampler = torchaudio.transforms.Resample(orig_freq=sr, new_freq=target_sr)
            wav = resampler(wav)
            
        # Ensure mono
        if wav.shape[0] > 1:
            wav = torch.mean(wav, dim=0, keepdim=True)
            
        # Move to device
        wav = wav.to(device)
        
        if language in ['hi', 'mr']:
            lang_code = language
        else:
            # Fallback for Indic model: default to Hindi
            # Note: We theoretically shouldn't be here for English if routing works
            lang_code = 'hi'
        
        with torch.no_grad():
            transcription = model(wav, lang_code, "ctc")
        
        return {
            'success': True,
            'text': str(transcription),
            'language': lang_code, 
            'language_probability': 1.0,
            'segments': []
        }
            
    except Exception as e:
        return {'success': False, 'error': str(e), 'text': '', 'language': language}

def transcribe_english(model_dict, audio_path: str, language: Optional[str] = None) -> Dict:
    """Transcribe using English Pipeline."""
    try:
        pipe = model_dict['model']
        
        # Pipeline handles loading and processing internally
        result = pipe(
            audio_path,
            chunk_length_s=30,
            batch_size=8,
            generate_kwargs={"language": "english"} # Force English
        )
        
        return {
            'success': True,
            'text': result['text'].strip(),
            'language': 'en',
            'language_probability': 1.0,
            'segments': []
        }
    except Exception as e:
        return {'success': False, 'error': str(e), 'text': '', 'language': language}

def transcribe_whisper(model_dict, audio_path: str, language: Optional[str] = None) -> Dict:
    """Transcribe using Distil-Whisper Pipeline."""
    try:
        pipe = model_dict['model']
        
        # Pipeline handles loading and processing internally
        result = pipe(
            audio_path,
            chunk_length_s=30,
            batch_size=8,
            generate_kwargs={"language": "english"} # Force English for distil-whisper
        )
        
        return {
            'success': True,
            'text': result['text'].strip(),
            'language': 'en',
            'language_probability': 1.0,
            'segments': []
        }
    except Exception as e:
        return {'success': False, 'error': str(e), 'text': '', 'language': language}

def get_audio_format_from_url(url: str) -> str:
    """Detect audio format from URL."""
    url_lower = url.lower()
    if '.mp3' in url_lower:
        return 'mp3'
    elif '.wav' in url_lower:
        return 'wav'
    elif '.ogg' in url_lower:
        return 'ogg'
    elif '.m4a' in url_lower:
        return 'm4a'
    else:
        return 'mp3'

def process_single_audio(models_collection, url: str, audio_type: str, language: Optional[str]) -> Dict:
    """Process a single audio file: download, extract, transcribe."""
    
    # Route based on language
    if language == 'en':
        model_dict = models_collection['english']
        transcribe_func = transcribe_english
    else:
        # Hindi, Marathi, or Unknown (assume Indic)
        model_dict = models_collection['indic']
        transcribe_func = transcribe_indic
    
    if not url or url.strip() == '':
        return {
            'url': url,
            'type': audio_type,
            'status': 'skipped',
            'reason': 'No URL provided',
            'transcription': '',
            'duration_seconds': 0,
            'duration_formatted': ''
        }
    
    print(f"  Processing {audio_type}: {url[:80]}...")
    
    # Download audio
    audio_bytes = download_audio(url)
    if audio_bytes is None:
        return {
            'url': url,
            'type': audio_type,
            'status': 'failed',
            'reason': 'Download failed',
            'transcription': '',
            'duration_seconds': 0,
            'duration_formatted': ''
        }
    
    # Extract audio content (full or truncated based on config)
    audio_format = get_audio_format_from_url(url)
    audio_path, duration_seconds = extract_audio_content(audio_bytes, audio_format)
    duration_formatted = format_duration(duration_seconds) if duration_seconds > 0 else ''
    
    if audio_path is None:
        return {
            'url': url,
            'type': audio_type,
            'status': 'failed',
            'reason': 'Audio processing failed',
            'transcription': '',
            'duration_seconds': 0,
            'duration_formatted': ''
        }
    
    try:
        # Transcribe
        result = transcribe_func(model_dict, audio_path, language)
        
        if result['success']:
            return {
                'url': url,
                'type': audio_type,
                'status': 'success',
                'transcription': result['text'],
                'detected_language': result['language'],
                'language_probability': result.get('language_probability', 0),
                'duration_seconds': duration_seconds,
                'duration_formatted': duration_formatted
            }
        else:
            return {
                'url': url,
                'type': audio_type,
                'status': 'failed',
                'reason': result.get('error', 'Transcription failed'),
                'transcription': '',
                'duration_seconds': duration_seconds,
                'duration_formatted': duration_formatted
            }
    finally:
        # Clean up temp file
        if audio_path and os.path.exists(audio_path):
            os.unlink(audio_path)

def generate_report(results: List[Dict], output_path: str):
    """Generate a comprehensive report in text format."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 100 + "\n")
        f.write("AUDIO TRANSCRIPTION REPORT (Hybrid: Whisper + IndicConformer)\n")
        f.write("=" * 100 + "\n\n")
        f.write(f"Generated: {timestamp}\n")
        f.write(f"Total Records Processed: {len(results)}\n\n")
        
        # Summary statistics
        total_audios = 0
        successful = 0
        failed = 0
        skipped = 0
        
        for record in results:
            for audio in record.get('audios', []):
                total_audios += 1
                if audio['status'] == 'success':
                    successful += 1
                elif audio['status'] == 'failed':
                    failed += 1
                else:
                    skipped += 1
        
        f.write("-" * 50 + "\n")
        f.write("SUMMARY\n")
        f.write("-" * 50 + "\n")
        f.write(f"Total Audio Files: {total_audios}\n")
        f.write(f"✓ Successfully Transcribed: {successful}\n")
        f.write(f"✗ Failed: {failed}\n")
        f.write(f"○ Skipped (No URL): {skipped}\n\n")
        
        # Detailed results
        f.write("=" * 100 + "\n")
        f.write("DETAILED TRANSCRIPTIONS\n")
        f.write("=" * 100 + "\n\n")
        
        for record in results:
            f.write("-" * 80 + "\n")
            f.write(f"Record ID: {record['id']}\n")
            f.write(f"Subject: {record.get('subject_name', 'N/A')} (ID: {record['subjectId']})\n")
            f.write(f"QSet: {record['qset']}\n")
            f.write(f"Department ID: {record.get('departmentId', 'N/A')}\n")
            f.write(f"Target Language: {record.get('language', 'Auto-detect')}\n")
            f.write("-" * 80 + "\n\n")
            
            for audio in record.get('audios', []):
                f.write(f"  [{audio['type'].upper()}]\n")
                f.write(f"  URL: {audio.get('url', 'N/A')}\n")
                f.write(f"  Duration: {audio.get('duration_formatted', 'N/A')} ({audio.get('duration_seconds', 0):.1f}s)\n")
                f.write(f"  Status: {audio['status'].upper()}\n")
                
                if audio['status'] == 'success':
                    f.write(f"  Detected Language: {audio.get('detected_language', 'N/A')}")
                    prob = audio.get('language_probability', 0)
                    if prob:
                        f.write(f" (confidence: {prob:.2%})\n")
                    else:
                        f.write("\n")
                    f.write(f"  Transcription:\n")
                    f.write(f"  {'-' * 40}\n")
                    text = audio['transcription']
                    for i in range(0, len(text), 80):
                        f.write(f"  {text[i:i+80]}\n")
                    f.write(f"  {'-' * 40}\n")
                elif audio['status'] == 'failed':
                    f.write(f"  Reason: {audio.get('reason', 'Unknown error')}\n")
                else:
                    f.write(f"  Reason: {audio.get('reason', 'No URL provided')}\n")
                
                f.write("\n")
            
            f.write("\n")
        
        f.write("=" * 100 + "\n")
        f.write("END OF REPORT\n")
        f.write("=" * 100 + "\n")
    
    print(f"✓ Report saved to: {output_path}")

def generate_json_report(results: List[Dict], output_path: str):
    """Generate a JSON report for programmatic access."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    report = {
        'generated_at': timestamp,
        'model': 'Hybrid (English: Distil-Whisper-v2 | Indic: AI4Bharat-Conformer)',
        'total_records': len(results),
        'results': results
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    
    print(f"✓ JSON report saved to: {output_path}")

def generate_excel_report(records_with_transcriptions: List[Dict], output_path: str):
    """Generate Excel report with all audiodb fields + transcription columns."""
    if not OPENPYXL_AVAILABLE:
        print("⚠ openpyxl not installed. Installing now...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        print("✓ openpyxl installed. Please restart the script to generate Excel report.")
        return
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Transcriptions"
    
    # Define headers (all audiodb fields + transcription columns)
    headers = [
        'id', 'subjectId', 'qset', 'departmentId',
        'code_a', 'code_b', 'code_t',
        'audio1', 'audio2', 'testaudio',
        'passage1', 'passage2',
        'textPassageA', 'textPassageB',
        'subject_name', 'subject_name_short',
        'passage1_duration', 'passage1_transcription', 'passage1_language', 'passage1_confidence',
        'passage2_duration', 'passage2_transcription', 'passage2_language', 'passage2_confidence',
        'testaudio_duration', 'testaudio_transcription', 'testaudio_language', 'testaudio_confidence'
    ]
    
    # Write headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write data rows
    row_idx = 2
    for record_data in records_with_transcriptions:
        # Get original database fields
        record = record_data['original_record']
        
        # Get transcriptions
        passage1_dur = ''
        passage1_trans = ''
        passage1_lang = ''
        passage1_conf = ''
        passage2_dur = ''
        passage2_trans = ''
        passage2_lang = ''
        passage2_conf = ''
        testaudio_dur = ''
        testaudio_trans = ''
        testaudio_lang = ''
        testaudio_conf = ''
        
        for audio in record_data.get('audios', []):
            if audio['type'] == 'passage1':
                passage1_dur = audio.get('duration_formatted', '')
                if audio['status'] == 'success':
                    passage1_trans = audio.get('transcription', '')
                    passage1_lang = audio.get('detected_language', '')
                    passage1_conf = f"{audio.get('language_probability', 0):.2%}"
            elif audio['type'] == 'passage2':
                passage2_dur = audio.get('duration_formatted', '')
                if audio['status'] == 'success':
                    passage2_trans = audio.get('transcription', '')
                    passage2_lang = audio.get('detected_language', '')
                    passage2_conf = f"{audio.get('language_probability', 0):.2%}"
            elif audio['type'] == 'testaudio':
                testaudio_dur = audio.get('duration_formatted', '')
                if audio['status'] == 'success':
                    testaudio_trans = audio.get('transcription', '')
                    testaudio_lang = audio.get('detected_language', '')
                    testaudio_conf = f"{audio.get('language_probability', 0):.2%}"
        
        # Write row data
        row_data = [
            record.get('id'),
            record.get('subjectId'),
            record.get('qset'),
            record.get('departmentId'),
            record.get('code_a'),
            record.get('code_b'),
            record.get('code_t'),
            record.get('audio1'),
            record.get('audio2'),
            record.get('testaudio'),
            record.get('passage1'),
            record.get('passage2'),
            record.get('textPassageA'),
            record.get('textPassageB'),
            record.get('subject_name'),
            record.get('subject_name_short'),
            passage1_dur,
            passage1_trans,
            passage1_lang,
            passage1_conf,
            passage2_dur,
            passage2_trans,
            passage2_lang,
            passage2_conf,
            testaudio_dur,
            testaudio_trans,
            testaudio_lang,
            testaudio_conf
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        row_idx += 1
    
    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    # Limit to reasonable width
                    cell_length = min(len(str(cell.value)), 50)
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max(12, max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel report saved to: {output_path}")

def main():
    print("\n" + "=" * 60)
    print("AUDIO TRANSCRIPTION REPORT GENERATOR V2")
    print("Hybrid Mode Activated:")
    print("1. English       -> Distil-Whisper Large-v2 (Fast, PyTorch)")
    print("2. Hindi/Marathi -> AI4Bharat Indic-Conformer-600M")
    print("=" * 60 + "\n")
    
    # Check if transformers is installed
    if not TRANSFORMERS_AVAILABLE:
        install_dependencies()
        print("Please restart the script after installation.")
        return
    
    # Check GPU availability
    has_gpu = check_gpu_availability()
    
    # Load Models
    indic_model = load_indic_model()
    english_model = load_english_model()
    
    models_collection = {
        'indic': indic_model,
        'english': english_model
    }
    
    # Connect to database
    connection = connect_to_database()
    
    try:
        # Fetch audio records
        records = fetch_audio_records(connection, department_id=TARGET_DEPARTMENT_ID)
        
        if len(records) == 0:
            print("No audio records found in the database.")
            return
        
        # Process each record
        results = []
        for i, record in enumerate(records, 1):
            print(f"\n[{i}/{len(records)}] Processing Record ID: {record['id']}")
            print(f"  Subject: {record.get('subject_name', 'Unknown')} (ID: {record['subjectId']})")
            print(f"  QSet: {record['qset']}")
            
            # Detect language from subject name
            language = detect_language_from_subject(record.get('subject_name'))
            lang_display = language if language else 'Auto-detect'
            print(f"  Language: {lang_display}")
            
            # Process each audio file
            audios = []
            
            # Passage 1
            if record.get('passage1'):
                audio_result = process_single_audio(models_collection, record['passage1'], 'passage1', language)
                audios.append(audio_result)
                dur_str = f" [{audio_result.get('duration_formatted', '')}]" if audio_result.get('duration_formatted') else ''
                if audio_result['status'] == 'success':
                    print(f"    ✓ Passage1{dur_str} transcribed ({len(audio_result['transcription'])} chars)")
                    print(f"      Text: {audio_result['transcription'][:100]}..." if len(audio_result['transcription']) > 100 else f"      Text: {audio_result['transcription']}")
                else:
                    print(f"    ✗ Passage1{dur_str} Failed: {audio_result.get('reason', 'Unknown error')}")
            
            # Passage 2
            if record.get('passage2'):
                audio_result = process_single_audio(models_collection, record['passage2'], 'passage2', language)
                audios.append(audio_result)
                dur_str = f" [{audio_result.get('duration_formatted', '')}]" if audio_result.get('duration_formatted') else ''
                if audio_result['status'] == 'success':
                    print(f"    ✓ Passage2{dur_str} transcribed ({len(audio_result['transcription'])} chars)")
                    print(f"      Text: {audio_result['transcription'][:100]}..." if len(audio_result['transcription']) > 100 else f"      Text: {audio_result['transcription']}")
                else:
                    print(f"    ✗ Passage2{dur_str} Failed: {audio_result.get('reason', 'Unknown error')}")
            
            # Test Audio
            if record.get('testaudio'):
                audio_result = process_single_audio(models_collection, record['testaudio'], 'testaudio', language)
                audios.append(audio_result)
                dur_str = f" [{audio_result.get('duration_formatted', '')}]" if audio_result.get('duration_formatted') else ''
                if audio_result['status'] == 'success':
                    print(f"    ✓ TestAudio{dur_str} transcribed ({len(audio_result['transcription'])} chars)")
                    print(f"      Text: {audio_result['transcription'][:100]}..." if len(audio_result['transcription']) > 100 else f"      Text: {audio_result['transcription']}")
                else:
                    print(f"    ✗ TestAudio{dur_str} Failed: {audio_result.get('reason', 'Unknown error')}")
            
            results.append({
                'id': record['id'],
                'subjectId': record['subjectId'],
                'qset': record['qset'],
                'departmentId': record.get('departmentId'),
                'subject_name': record.get('subject_name'),
                'subject_name_short': record.get('subject_name_short'),
                'language': language,
                'audios': audios,
                'original_record': record  # Store original database record for Excel export
            })
        
        # Generate reports
        report_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Build suffix for output filenames
        dept_suffix = f'_dept{TARGET_DEPARTMENT_ID}' if TARGET_DEPARTMENT_ID else ''

        # Excel report (primary output)
        excel_report_path = os.path.join(report_dir, f'audio_transcription_report{dept_suffix}.xlsx')
        generate_excel_report(results, excel_report_path)

        # Text report
        text_report_path = os.path.join(report_dir, f'audio_transcription_report_v2{dept_suffix}.txt')
        generate_report(results, text_report_path)

        # JSON report
        json_report_path = os.path.join(report_dir, f'audio_transcription_report_v2{dept_suffix}.json')
        generate_json_report(results, json_report_path)
        
        print("\n" + "=" * 60)
        print("PROCESSING COMPLETE!")
        print("=" * 60)
        print(f"\nReports generated:")
        print(f"  1. {excel_report_path} (Primary)")
        print(f"  2. {text_report_path}")
        print(f"  3. {json_report_path}")
        
    finally:
        connection.close()
        print("\n✓ Database connection closed")

if __name__ == "__main__":
    main()
