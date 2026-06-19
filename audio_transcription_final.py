"""
Audio Transcription Report Generator - Whisper Large-v3-Turbo
==============================================================
Uses faster-whisper with large-v3-turbo model
8× faster than large-v3, excellent for Hindi/Marathi/English
Generates Excel report with audiodb fields + transcriptions
"""

import os
import sys
import mysql.connector
from datetime import datetime
from pydub import AudioSegment
from typing import Optional, Dict, List
import io
import warnings
import tempfile
import requests
warnings.filterwarnings("ignore")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

try:
    from faster_whisper import WhisperModel
except ImportError:
    print("Installing faster-whisper...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "faster-whisper"])
    from faster_whisper import WhisperModel

try:
    import torch
    import librosa
    from transformers import Wav2Vec2ForCTC, Wav2Vec2Processor
except ImportError:
    print("Installing transformers torch librosa...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "transformers", "torch", "librosa"])
    import torch
    import librosa
    from transformers import Wav2Vec2ForCTC, Wav2Vec2Processor

# Database configuration
DB_CONFIG = {
    'host': '13.204.48.33',
    'port': 3306,
    'user': 'root',
    'password': 'tanuj1221',
    'database': 'dec25',
    'charset': 'utf8mb4'
}


SAMPLE_RATE = 16000

def check_gpu():
    """Check GPU availability."""
    if torch.cuda.is_available():
        gpu_name = torch.cuda.get_device_name(0)
        print(f"✓ GPU Available: {gpu_name}")
        return "cuda"
    print("✗ Using CPU (slower)")
    return "cpu"

def load_models(device_type):
    """Load both Whisper (English/Default) and Wav2Vec2 (Hindi/Marathi) models."""
    models = {}
    
    # 1. Load Faster Whisper
    compute_type = "float16" if device_type == "cuda" else "int8"
    print(f"\nLoading Whisper large-v3-turbo on {device_type.upper()} (for English)...")
    try:
        models['whisper'] = WhisperModel(
            "large-v3-turbo",
            device=device_type,
            compute_type=compute_type,
            num_workers=4
        )
        print(f"✓ Whisper loaded")
    except Exception as e:
        print(f"✗ Failed to load Whisper: {e}")
        models['whisper'] = None

    # 2. Load Wav2Vec2 for Hindi/Marathi
    print(f"Loading AI4Bharat IndicWav2Vec on {device_type.upper()}...")
    try:
        # Using AI4Bharat model which is generally more accurate than the previous one
        # Note: This is the Transformers-compatible version of their work
        model_name = "ai4bharat/indicwav2vec_v1_ada_rom_ft"
        processor = Wav2Vec2Processor.from_pretrained(model_name)
        model = Wav2Vec2ForCTC.from_pretrained(model_name)
        model.to(device_type)
        
        models['hindi'] = {
            'model': model,
            'processor': processor,
            'device': device_type
        }
        print(f"✓ AI4Bharat IndicWav2Vec loaded")
    except Exception as e:
        print(f"✗ Failed to load Wav2Vec2: {e}")
        models['hindi'] = None
        
    return models

def connect_db():
    """Connect to MySQL database."""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        print(f"✓ Connected to database: {DB_CONFIG['database']}")
        return conn
    except Exception as e:
        print(f"✗ Database error: {e}")
        sys.exit(1)

def fetch_records(conn):
    """Fetch audio records from database."""
    cursor = conn.cursor(dictionary=True)
    query = """
    SELECT 
        a.id, a.subjectId, a.qset, a.departmentId,
        a.code_a, a.code_b, a.code_t,
        a.audio1, a.audio2, a.testaudio,
        a.passage1, a.passage2,
        a.textPassageA, a.textPassageB,
        s.subject_name
    FROM audiodb a
    LEFT JOIN (
        SELECT subjectId, MAX(subject_name) as subject_name 
        FROM subjectsdb 
        GROUP BY subjectId
    ) s ON a.subjectId = s.subjectId
    ORDER BY a.subjectId, a.qset
    """
    cursor.execute(query)
    records = cursor.fetchall()
    cursor.close()
    print(f"✓ Fetched {len(records)} audio records (Joined with Subjects)\n")
    return records

def detect_language(subject_name):
    """Detect language from subject name."""
    if not subject_name:
        return None
    subject_lower = subject_name.lower()
    if 'hindi' in subject_lower:
        return 'hi'
    elif 'marathi' in subject_lower:
        return 'mr'
    elif 'english' in subject_lower:
        return 'en'
    return None

def download_audio(url):
    """Download audio from URL."""
    if not url or not url.strip():
       return None
    try:
        response = requests.get(url, timeout=60, stream=True)
        response.raise_for_status()
        return response.content
    except Exception as e:
        print(f"  ⚠ Download failed: {str(e)[:50]}")
        return None

def prepare_audio(audio_bytes, format="mp3"):
    """Prepare audio for transcription (convert to 16kHz mono WAV)."""
    try:
        audio = AudioSegment.from_file(io.BytesIO(audio_bytes), format=format)
        # Convert to 16kHz mono, keep full duration
        audio = audio.set_channels(1).set_frame_rate(16000)
        
        temp_file = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
        audio.export(temp_file.name, format="wav")
        temp_file.close()
        return temp_file.name
    except Exception as e:
        print(f"  ⚠ Audio processing failed: {str(e)[:50]}")
        return None

def transcribe_whisper(model, audio_path, language=None):
    """Transcribe using Whisper."""
    try:
        segments, info = model.transcribe(
            audio_path,
            language=language,
            beam_size=5,
            vad_filter=True
        )
        text = " ".join([seg.text for seg in segments]).strip()
        return {
            'success': True,
            'text': text,
            'language': info.language,
            'probability': info.language_probability
        }
    except Exception as e:
        return {'success': False, 'error': str(e), 'text': '', 'language': language}

def transcribe_wav2vec2(model_dict, audio_path):
    """Transcribe using Wav2Vec2 for Hindi/Marathi."""
    try:
        model = model_dict['model']
        processor = model_dict['processor']
        device = model_dict['device']
        
        # Load audio using librosa (resample to 16k)
        audio_input, sr = librosa.load(audio_path, sr=16000)
        
        # Tokenize
        inputs = processor(audio_input, sampling_rate=16000, return_tensors="pt", padding=True)
        
        # Move to device
        inputs = inputs.to(device)
        
        # Inference
        with torch.no_grad():
            logits = model(inputs.input_values).logits
            
        # Decode
        predicted_ids = torch.argmax(logits, dim=-1)
        transcription = processor.batch_decode(predicted_ids)[0]
        
        return {
            'success': True,
            'text': transcription,
            'language': 'hi/mr', # grouped
            'probability': 1.0
        }
    except Exception as e:
        return {'success': False, 'error': str(e), 'text': '', 'language': 'hi/mr'}

def transcribe(models, audio_path, language=None):
    """Dispatch transcription based on language."""
    # Logic: If language is English or Unknown -> Whisper
    # If language is Hindi or Marathi -> Wav2Vec2 (if available), else Whisper
    
    use_hindi_model = False
    if language in ['hi', 'mr'] and models.get('hindi'):
        use_hindi_model = True
        
    if use_hindi_model:
        return transcribe_wav2vec2(models['hindi'], audio_path)
    else:
        if models.get('whisper'):
            # For whisper, if language is 'hi' or 'mr' but we are here, it means hindi model failed loading?
            # Or if language is None/en.
            # Convert 'hi'/'mr' to None if we are falling back to Whisper so it auto-detects or explicitly pass it?
            # Whisper handles 'hi'/'mr' well too, but user specifically requested the other model.
            return transcribe_whisper(models['whisper'], audio_path, language)
        else:
            return {'success': False, 'error': "No suitable model loaded", 'text': '', 'language': language}

def process_audio(model, url, audio_type, language):
    """Process single audio: download, extract, transcribe."""
    if not url or not url.strip():
        return {
            'type': audio_type,
            'status': 'skipped',
            'transcription': ''
        }
    
    print(f"  Processing {audio_type}...")
    
    # Download
    audio_bytes = download_audio(url)
    if not audio_bytes:
        return {'type': audio_type, 'status': 'failed', 'transcription': ''}
    
    # Prepare audio (convert/resample)
    format = 'mp3' if '.mp3' in url.lower() else 'wav'
    audio_path = prepare_audio(audio_bytes, format)
    if not audio_path:
        return {'type': audio_type, 'status': 'failed', 'transcription': ''}
    
    try:
        # Transcribe
        result = transcribe(model, audio_path, language)
        
        if result['success']:
            print(f"    ✓ {audio_type} transcribed ({len(result['text'])} chars)")
            return {
                'type': audio_type,
                'status': 'success',
                'transcription': result['text'],
                'language': result['language'],
                'probability': result.get('probability', 0)
            }
        else:
            return {'type': audio_type, 'status': 'failed', 'transcription': ''}
    finally:
        if audio_path and os.path.exists(audio_path):
            os.unlink(audio_path)

def generate_excel(results, output_path):
    """Generate Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Transcriptions"
    
    # Headers
    headers = [
        'id', 'subjectId', 'qset', 'departmentId',
        'code_a', 'code_b', 'code_t',
        'audio1', 'audio2', 'testaudio',
        'passage1', 'passage2',
        'textPassageA', 'textPassageB',
        'subject_name', 'subject_name_short',
        'passage1_transcription', 'passage1_language',
        'passage2_transcription', 'passage2_language',
        'testaudio_transcription', 'testaudio_language'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data rows
    row_idx = 2
    for rec in results:
        record = rec['original_record']
        
        # Extract transcriptions
        p1_trans = p1_lang = p2_trans = p2_lang = t_trans = t_lang = ''
        
        for audio in rec.get('audios', []):
            if audio['type'] == 'passage1' and audio['status'] == 'success':
                p1_trans = audio.get('transcription', '')
                p1_lang = audio.get('language', '')
            elif audio['type'] == 'passage2' and audio['status'] == 'success':
                p2_trans = audio.get('transcription', '')
                p2_lang = audio.get('language', '')
            elif audio['type'] == 'testaudio' and audio['status'] == 'success':
                t_trans = audio.get('transcription', '')
                t_lang = audio.get('language', '')
        
        # Row data
        row_data = [
            record.get('id'), record.get('subjectId'), record.get('qset'),
            record.get('departmentId'), record.get('code_a'), record.get('code_b'),
            record.get('code_t'), record.get('audio1'), record.get('audio2'),
            record.get('testaudio'), record.get('passage1'), record.get('passage2'),
            record.get('textPassageA'), record.get('textPassageB'),
            record.get('subject_name'), record.get('subject_name_short'),
            p1_trans, p1_lang, p2_trans, p2_lang, t_trans, t_lang
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        row_idx += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
            except:
                pass
        ws.column_dimensions[column_letter].width = max(12, max_length + 2)
    
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"\n✓ Excel report saved: {output_path}")

def main():
    print("\n" + "=" * 70)
    print("AUDIO TRANSCRIPTION REPORT GENERATOR")
    print("Using Whisper Large-v3-Turbo (8× faster, best for Indian languages)")
    print("=" * 70 + "\n")
    
    # Check GPU
    device_type = check_gpu()
    
    # Load models
    models = load_models(device_type)
    
    # Connect to database
    conn = connect_db()
    
    try:
        # Fetch records
        records = fetch_records(conn)
        
        if not records:
            print("No audio records found.")
            return
        
        # Process each record
        results = []
        for i, record in enumerate(records, 1):
            print(f"[{i}/{len(records)}] Record ID: {record['id']}")
            print(f"  Subject: {record.get('subject_name', 'Unknown')}")
            
            language = detect_language(record.get('subject_name'))
            
            audios = []
            
            # Process passage1
            if record.get('passage1'):
                audios.append(process_audio(models, record['passage1'], 'passage1', language))
            
            # Process passage2
            if record.get('passage2'):
                audios.append(process_audio(models, record['passage2'], 'passage2', language))
            
            # Process testaudio
            if record.get('testaudio'):
                audios.append(process_audio(models, record['testaudio'], 'testaudio', language))
            
            results.append({
                'original_record': record,
                'audios': audios
            })
            print()
        
        # Generate Excel report
        report_path = os.path.join(os.path.dirname(__file__), 'audio_transcription_report.xlsx')
        generate_excel(results, report_path)
        
        # Summary
        total = sum(len(r['audios']) for r in results)
        success = sum(1 for r in results for a in r['audios'] if a['status'] == 'success')
        
        print("\n" + "=" * 70)
        print("PROCESSING COMPLETE!")
        print("=" * 70)
        print(f"\nTotal audio files: {total}")
        print(f"Successfully transcribed: {success}")
        print(f"Failed: {total - success}")
        
    finally:
        conn.close()
        print("\n✓ Database connection closed\n")

if __name__ == "__main__":
    main()
