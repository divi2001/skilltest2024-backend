"""
Audio Transcription Report Generator - IndicWhisper
====================================================
Uses AI4Bharat's IndicWhisper for best Hindi/Marathi/English accuracy
This model is specifically trained on Indian languages
Generates Excel report with audiodb fields + transcriptions
"""

import os
import sys
import json
import mysql.connector
from datetime import datetime
from pydub import AudioSegment
from typing import Optional, Dict, List
import io
import warnings
import tempfile
import requests
import torch
warnings.filterwarnings("ignore")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from transformers import pipeline
    import torchaudio
    TRANSFORMERS_AVAILABLE = True
except ImportError:
    TRANSFORMERS_AVAILABLE = False
    print("transformers not installed. Will install automatically...")

# Database configuration
DB_CONFIG = {
    'host': '13.204.48.33',
    'port': 3306,
    'user': 'root',
    'password': 'tanuj1221',
    'database': 'dec25',
    'charset': 'utf8mb4'
}

# Audio processing configuration
AUDIO_DURATION_SECONDS = 30
SAMPLE_RATE = 16000

# Use Whisper Large-v3-Turbo - Best for Indian languages, no authentication needed
# This model is 8× faster than large-v3 and has better Hindi/Marathi support
WHISPER_MODEL = "openai/whisper-large-v3-turbo"

def install_dependencies():
    """Install required packages."""
    import subprocess
    packages = ["transformers", "accelerate", "torchaudio"]
    for package in packages:
        print(f"Installing {package}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    print("✓ All dependencies installed successfully")

def check_gpu_availability():
    """Check if GPU is available for processing."""
    if torch.cuda.is_available():
        gpu_name = torch.cuda.get_device_name(0)
        gpu_memory = torch.cuda.get_device_properties(0).total_memory / (1024**3)
        print(f"✓ GPU Available: {gpu_name} ({gpu_memory:.2f} GB)")
        return True
    else:
        print("✗ GPU not available. Using CPU (slower).")
        return False

def load_indicwhisper_model():
    """Load IndicWhisper model with GPU support."""
    device = "cuda" if torch.cuda.is_available() else "cpu"
    
    print(f"Loading IndicWhisper model on {device.upper()}...")
    print("Model: ai4bharat/indicwhisper-medium-hi-en")
    print("Optimized for: Hindi, English, and other Indian languages")
    
    # Load the ASR pipeline
    asr_pipeline = pipeline(
        "automatic-speech-recognition",
        model=INDICWHISPER_MODEL,
        device=0 if device == "cuda" else -1,  # 0 for GPU, -1 for CPU
        torch_dtype=torch.float16 if device == "cuda" else torch.float32
    )
    
    print(f"✓ Model loaded successfully on {device.upper()}")
    return asr_pipeline

def connect_to_database():
    """Establish connection to MySQL database."""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        print(f"✓ Connected to database: {DB_CONFIG['database']}")
        return connection
    except mysql.connector.Error as err:
        print(f"✗ Database connection failed: {err}")
        sys.exit(1)

def fetch_audio_records(connection) -> List[Dict]:
    """Fetch all audio records from audiodb table."""
    cursor = connection.cursor(dictionary=True)
    
    query = """
    SELECT 
        a.id,
        a.subjectId,
        a.qset,
        a.departmentId,
        a.code_a,
        a.code_b,
        a.code_t,
        a.audio1,
        a.audio2,
        a.testaudio,
        a.passage1,
        a.passage2,
        a.textPassageA,
        a.textPassageB,
        s.subject_name,
        s.subject_name_short
    FROM audiodb a
    LEFT JOIN subjectsdb s ON a.subjectId = s.subjectId
    ORDER BY a.subjectId, a.qset
    """
    cursor.execute(query)
    records = cursor.fetchall()
    cursor.close()
    
    print(f"✓ Fetched {len(records)} audio records from database")
    return records

def detect_language_from_subject(subject_name: Optional[str]) -> str:
    """Detect language based on subject name."""
    if subject_name is None:
        return None
    
    subject_lower = subject_name.lower()
    
    if 'hindi' in subject_lower or 'हिंदी' in subject_lower:
        return 'hi'
    elif 'marathi' in subject_lower or 'मराठी' in subject_lower:
        return 'mr'
    elif 'english' in subject_lower:
        return 'en'
    else:
        return None  # Auto-detect

def download_audio(url: str, timeout: int = 60) -> Optional[bytes]:
    """Download audio file from URL."""
    if not url or url.strip() == '':
        return None
    
    try:
        response = requests.get(url, timeout=timeout, stream=True)
        response.raise_for_status()
        return response.content
    except requests.RequestException as e:
        print(f"  ⚠ Failed to download audio: {e}")
        return None

def extract_first_30_seconds(audio_bytes: bytes, format: str = "mp3") -> Optional[str]:
    """Extract first 30 seconds from audio file and save to temp file."""
    try:
        # Load audio from bytes
        audio = AudioSegment.from_file(io.BytesIO(audio_bytes), format=format)
        
        # Extract first 30 seconds
        duration_ms = AUDIO_DURATION_SECONDS * 1000
        first_30_sec = audio[:duration_ms]
        
        # Convert to mono and set sample rate
        first_30_sec = first_30_sec.set_channels(1).set_frame_rate(SAMPLE_RATE)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
        first_30_sec.export(temp_file.name, format="wav")
        temp_file.close()
        
        return temp_file.name
    except Exception as e:
        print(f"  ⚠ Failed to process audio: {e}")
        return None

def transcribe_audio(pipeline_model, audio_path: str, language: Optional[str] = None) -> Dict:
    """Transcribe audio using IndicWhisper model."""
    try:
        # Load audio with torchaudio
        waveform, sample_rate = torchaudio.load(audio_path)
        
        # Resample if needed
        if sample_rate != SAMPLE_RATE:
            resampler = torchaudio.transforms.Resample(sample_rate, SAMPLE_RATE)
            waveform = resampler(waveform)
        
        # Convert to mono if stereo
        if waveform.shape[0] > 1:
            waveform = torch.mean(waveform, dim=0, keepdim=True)
        
        # Convert to numpy array
        audio_array = waveform.squeeze().numpy()
        
        # Transcribe
        # IndicWhisper supports language hints
        generate_kwargs = {}
        if language:
            # Map language codes to IndicWhisper format
            lang_map = {'hi': 'hindi', 'mr': 'marathi', 'en': 'english'}
            if language in lang_map:
                generate_kwargs['language'] = lang_map[language]
        
        result = pipeline_model(
            audio_array,
            generate_kwargs=generate_kwargs,
            return_timestamps=False
        )
        
        transcription = result['text'].strip()
        detected_lang = language if language else 'auto'
        
        return {
            'success': True,
            'text': transcription,
            'language': detected_lang,
            'language_probability': 1.0  # IndicWhisper doesn't provide this
        }
            
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'text': '',
            'language': language
        }

def get_audio_format_from_url(url: str) -> str:
    """Detect audio format from URL."""
    url_lower = url.lower()
    if '.mp3' in url_lower:
        return 'mp3'
    elif '.wav' in url_lower:
        return 'wav'
    elif '.ogg' in url_lower:
        return 'ogg'
    elif '.m4a' in url_lower:
        return 'm4a'
    else:
        return 'mp3'

def process_single_audio(model, url: str, audio_type: str, language: Optional[str]) -> Dict:
    """Process a single audio file: download, extract 30s, transcribe."""
    if not url or url.strip() == '':
        return {
            'url': url,
            'type': audio_type,
            'status': 'skipped',
            'reason': 'No URL provided',
            'transcription': ''
        }
    
    print(f"  Processing {audio_type}: {url[:80]}...")
    
    # Download audio
    audio_bytes = download_audio(url)
    if audio_bytes is None:
        return {
            'url': url,
            'type': audio_type,
            'status': 'failed',
            'reason': 'Download failed',
            'transcription': ''
        }
    
    # Extract first 30 seconds
    audio_format = get_audio_format_from_url(url)
    audio_path = extract_first_30_seconds(audio_bytes, audio_format)
    if audio_path is None:
        return {
            'url': url,
            'type': audio_type,
            'status': 'failed',
            'reason': 'Audio processing failed',
            'transcription': ''
        }
    
    try:
        # Transcribe
        result = transcribe_audio(model, audio_path, language)
        
        if result['success']:
            return {
                'url': url,
                'type': audio_type,
                'status': 'success',
                'transcription': result['text'],
                'detected_language': result['language'],
                'language_probability': result.get('language_probability', 1.0)
            }
        else:
            return {
                'url': url,
                'type': audio_type,
                'status': 'failed',
                'reason': result.get('error', 'Transcription failed'),
                'transcription': ''
            }
    finally:
        # Clean up temp file
        if audio_path and os.path.exists(audio_path):
            os.unlink(audio_path)

def generate_excel_report(records_with_transcriptions: List[Dict], output_path: str):
    """Generate Excel report with all audiodb fields + transcription columns."""
    if not OPENPYXL_AVAILABLE:
        print("⚠ openpyxl not installed. Installing now...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        print("✓ openpyxl installed. Please restart the script.")
        return
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Transcriptions"
    
    # Define headers
    headers = [
        'id', 'subjectId', 'qset', 'departmentId',
        'code_a', 'code_b', 'code_t',
        'audio1', 'audio2', 'testaudio',
        'passage1', 'passage2',
        'textPassageA', 'textPassageB',
        'subject_name', 'subject_name_short',
        'passage1_transcription', 'passage1_language', 'passage1_status',
        'passage2_transcription', 'passage2_language', 'passage2_status',
        'testaudio_transcription', 'testaudio_language', 'testaudio_status'
    ]
    
    # Write headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write data rows
    row_idx = 2
    for record_data in records_with_transcriptions:
        record = record_data['original_record']
        
        # Get transcriptions
        passage1_trans = ''
        passage1_lang = ''
        passage1_status = ''
        passage2_trans = ''
        passage2_lang = ''
        passage2_status = ''
        testaudio_trans = ''
        testaudio_lang = ''
        testaudio_status = ''
        
        for audio in record_data.get('audios', []):
            if audio['type'] == 'passage1':
                passage1_status = audio['status']
                if audio['status'] == 'success':
                    passage1_trans = audio.get('transcription', '')
                    passage1_lang = audio.get('detected_language', '')
            elif audio['type'] == 'passage2':
                passage2_status = audio['status']
                if audio['status'] == 'success':
                    passage2_trans = audio.get('transcription', '')
                    passage2_lang = audio.get('detected_language', '')
            elif audio['type'] == 'testaudio':
                testaudio_status = audio['status']
                if audio['status'] == 'success':
                    testaudio_trans = audio.get('transcription', '')
                    testaudio_lang = audio.get('detected_language', '')
        
        # Write row data
        row_data = [
            record.get('id'),
            record.get('subjectId'),
            record.get('qset'),
            record.get('departmentId'),
            record.get('code_a'),
            record.get('code_b'),
            record.get('code_t'),
            record.get('audio1'),
            record.get('audio2'),
            record.get('testaudio'),
            record.get('passage1'),
            record.get('passage2'),
            record.get('textPassageA'),
            record.get('textPassageB'),
            record.get('subject_name'),
            record.get('subject_name_short'),
            passage1_trans,
            passage1_lang,
            passage1_status,
            passage2_trans,
            passage2_lang,
            passage2_status,
            testaudio_trans,
            testaudio_lang,
            testaudio_status
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        row_idx += 1
    
    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = min(len(str(cell.value)), 50)
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max(12, max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel report saved to: {output_path}")

def main():
    print("\n" + "=" * 70)
    print("AUDIO TRANSCRIPTION REPORT GENERATOR")
    print("Using AI4Bharat's IndicWhisper")
    print("Optimized for Hindi, Marathi, and English")
    print("=" * 70 + "\n")
    
    # Check if transformers is installed
    if not TRANSFORMERS_AVAILABLE:
        install_dependencies()
        print("Please restart the script after installation.")
        return
    
    # Check GPU availability
    has_gpu = check_gpu_availability()
    
    # Load IndicWhisper model
    model = load_indicwhisper_model()
    
    # Connect to database
    connection = connect_to_database()
    
    try:
        # Fetch audio records
        records = fetch_audio_records(connection)
        
        if len(records) == 0:
            print("No audio records found in the database.")
            return
        
        # Process each record
        results = []
        for i, record in enumerate(records, 1):
            print(f"\n[{i}/{len(records)}] Processing Record ID: {record['id']}")
            print(f"  Subject: {record.get('subject_name', 'Unknown')} (ID: {record['subjectId']})")
            print(f"  QSet: {record['qset']}")
            
            # Detect language from subject name
            language = detect_language_from_subject(record.get('subject_name'))
            lang_display = language if language else 'Auto-detect'
            print(f"  Language: {lang_display}")
            
            # Process each audio file
            audios = []
            
            # Passage 1
            if record.get('passage1'):
                audio_result = process_single_audio(model, record['passage1'], 'passage1', language)
                audios.append(audio_result)
                if audio_result['status'] == 'success':
                    print(f"    ✓ Passage1 transcribed ({len(audio_result['transcription'])} chars)")
            
            # Passage 2
            if record.get('passage2'):
                audio_result = process_single_audio(model, record['passage2'], 'passage2', language)
                audios.append(audio_result)
                if audio_result['status'] == 'success':
                    print(f"    ✓ Passage2 transcribed ({len(audio_result['transcription'])} chars)")
            
            # Test Audio
            if record.get('testaudio'):
                audio_result = process_single_audio(model, record['testaudio'], 'testaudio', language)
                audios.append(audio_result)
                if audio_result['status'] == 'success':
                    print(f"    ✓ TestAudio transcribed ({len(audio_result['transcription'])} chars)")
            
            results.append({
                'id': record['id'],
                'subjectId': record['subjectId'],
                'qset': record['qset'],
                'departmentId': record.get('departmentId'),
                'subject_name': record.get('subject_name'),
                'subject_name_short': record.get('subject_name_short'),
                'language': language,
                'audios': audios,
                'original_record': record
            })
        
        # Generate Excel report
        report_dir = os.path.dirname(os.path.abspath(__file__))
        excel_report_path = os.path.join(report_dir, 'audio_transcription_report_indicwhisper.xlsx')
        generate_excel_report(results, excel_report_path)
        
        print("\n" + "=" * 70)
        print("PROCESSING COMPLETE!")
        print("=" * 70)
        print(f"\n✓ Excel report generated: {excel_report_path}")
        print(f"\nTotal records processed: {len(results)}")
        
        # Summary
        total_audios = sum(len(r['audios']) for r in results)
        successful = sum(1 for r in results for a in r['audios'] if a['status'] == 'success')
        print(f"Total audio files: {total_audios}")
        print(f"Successfully transcribed: {successful}")
        print(f"Failed: {total_audios - successful}")
        
    finally:
        connection.close()
        print("\n✓ Database connection closed")

if __name__ == "__main__":
    main()
